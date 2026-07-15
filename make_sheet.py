# -*- coding: utf-8 -*-
"""Tạo file Google Sheet đăng ký ca — Cơ sở 3."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

DOW   = ['T2','T3','T4','T5','T6','T7','CN']
PERS  = [('Sáng','m'), ('Chiều','a'), ('Tối','e')]
QUOTA = {
    'm': {'Pha chế':2, 'Phục vụ':3},
    'a': {'Pha chế':2, 'Phục vụ':2},
    'e': {'Pha chế':2, 'Phục vụ':3},
}
NDATA = 16          # số dòng nhân viên
HR    = 4           # dòng tiêu đề bảng
DR    = HR + 1      # dòng dữ liệu đầu
LR    = DR + NDATA - 1

# màu
BRAND   = '9A461C'
CREAM   = 'FBF5EC'
TAN_A   = 'F3E7D6'   # tint ngày lẻ
TAN_B   = 'FBF3E6'   # tint ngày chẵn
RED     = 'F6C6C0'
GREEN   = 'CDE9D5'
HEADTXT = '3A2A1E'
thin = Side(style='thin', color='E3D6C4')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Đăng ký'

# ---- tiêu đề & hướng dẫn ----
ws['A1'] = 'ĐĂNG KÝ CA LÀM — CƠ SỞ 3'
ws['A1'].font = Font(size=16, bold=True, color=BRAND)
ws['A2'] = ('Điền chữ  x  vào buổi bạn ĐI LÀM ĐƯỢC. Ai điền trước được ưu tiên. '
            'Ô chuyển ĐỎ = buổi đó đã đủ người → hãy chọn buổi/ngày khác. Ô XANH = đã nhận chỗ.')
ws['A2'].font = Font(size=10, italic=True, color='6B5B4C')
ws['A3'] = 'Định mức mỗi buổi:  Pha chế 2  ·  Phục vụ 2–3   (người giữ xe do chủ quán chọn trong app)'
ws['A3'].font = Font(size=10, bold=True, color='6B5B4C')

# ---- hàng tiêu đề bảng ----
hdr_fill = PatternFill('solid', fgColor=CREAM)
ws.cell(HR,1,'Tên').font = Font(bold=True, color=HEADTXT)
ws.cell(HR,2,'Vị trí').font = Font(bold=True, color=HEADTXT)
for c in (1,2):
    ws.cell(HR,c).fill = hdr_fill; ws.cell(HR,c).border = border
    ws.cell(HR,c).alignment = Alignment(horizontal='center', vertical='center')

col = 3
claim_cols = {}   # (d,pk) -> column index
for d,dow in enumerate(DOW):
    tint = TAN_A if d % 2 else TAN_B
    for plabel,pk in PERS:
        cell = ws.cell(HR, col, f'{dow} {plabel}')
        cell.font = Font(bold=True, size=10, color=HEADTXT)
        cell.fill = PatternFill('solid', fgColor=tint)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        claim_cols[(d,pk)] = col
        col += 1
last_col = col - 1

# ---- ô dữ liệu (viền + căn giữa) ----
for r in range(DR, LR+1):
    ws.cell(r,1).border = border
    ws.cell(r,2).border = border
    ws.cell(r,2).alignment = Alignment(horizontal='center')
    for cc in range(3, last_col+1):
        cell = ws.cell(r, cc)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        # tint nhạt theo ngày để dễ nhìn
        d = (cc-3)//3
        cell.fill = PatternFill('solid', fgColor=(TAN_A if d % 2 else TAN_B))

# ---- data validation ----
dv_role = DataValidation(type='list', formula1='"Pha chế,Phục vụ"', allow_blank=True)
ws.add_data_validation(dv_role); dv_role.add(f'B{DR}:B{LR}')
dv_x = DataValidation(type='list', formula1='"x"', allow_blank=True)
ws.add_data_validation(dv_x)
dv_x.add(f'{get_column_letter(3)}{DR}:{get_column_letter(last_col)}{LR}')

# ---- conditional formatting: đỏ khi vượt định mức (theo thứ tự điền), xanh khi hợp lệ ----
red_fill   = PatternFill('solid', fgColor=RED)
green_fill = PatternFill('solid', fgColor=GREEN)
for (d,pk), c in claim_cols.items():
    L = get_column_letter(c)
    q = QUOTA[pk]
    qexpr = (f'IF($B{DR}="Pha chế",{q["Pha chế"]},'
             f'IF($B{DR}="Phục vụ",{q["Phục vụ"]},99))')
    run = f'COUNTIFS($B${DR}:$B{DR},$B{DR},{L}${DR}:{L}{DR},"x")'
    rng = f'{L}{DR}:{L}{LR}'
    over  = f'AND({L}{DR}="x",{run}>{qexpr})'
    okok  = f'AND({L}{DR}="x",{run}<={qexpr})'
    ws.conditional_formatting.add(rng, FormulaRule(formula=[over], fill=red_fill, stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[okok], fill=green_fill))

# ---- kích thước cột / freeze ----
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 12
for cc in range(3, last_col+1):
    ws.column_dimensions[get_column_letter(cc)].width = 8.5
ws.row_dimensions[HR].height = 30
ws.freeze_panes = 'C5'

# =============== Sheet 2: Còn chỗ (đếm trực tiếp) ===============
ws2 = wb.create_sheet('Còn chỗ')
ws2['A1'] = 'CÒN CHỖ MỖI BUỔI (tự cập nhật)'
ws2['A1'].font = Font(size=14, bold=True, color=BRAND)
ws2['A2'] = 'Số đang có / định mức. Ô đỏ = đã đủ.'
ws2['A2'].font = Font(size=10, italic=True, color='6B5B4C')
heads = ['Ngày','Buổi','Pha chế','Phục vụ','Giữ xe']
for i,h in enumerate(heads,1):
    cc = ws2.cell(4,i,h); cc.font = Font(bold=True, color=HEADTXT)
    cc.fill = PatternFill('solid', fgColor=CREAM); cc.border = border
    cc.alignment = Alignment(horizontal='center')
roles = ['Pha chế','Phục vụ']
r2 = 5
red2 = PatternFill('solid', fgColor=RED)
for d,dow in enumerate(DOW):
    for plabel,pk in PERS:
        ws2.cell(r2,1,dow).alignment = Alignment(horizontal='center')
        ws2.cell(r2,2,plabel).alignment = Alignment(horizontal='center')
        for ri,role in enumerate(roles):
            c = claim_cols[(d,pk)]; L = get_column_letter(c)
            q = QUOTA[pk][role]
            f = (f'=COUNTIFS(\'Đăng ký\'!$B${DR}:$B${LR},"{role}",'
                 f'\'Đăng ký\'!{L}${DR}:{L}${LR},"x")&"/{q}"')
            cell = ws2.cell(r2, 3+ri, f)
            cell.alignment = Alignment(horizontal='center'); cell.border = border
            # đỏ khi đủ: count>=quota
            cnt = (f'COUNTIFS(\'Đăng ký\'!$B${DR}:$B${LR},"{role}",'
                   f'\'Đăng ký\'!{L}${DR}:{L}${LR},"x")')
            ws2.conditional_formatting.add(
                cell.coordinate, FormulaRule(formula=[f'{cnt}>={q}'], fill=red2))
        for i in (1,2):
            ws2.cell(r2,i).border = border
        r2 += 1
for i,w in enumerate([6,8,10,10,10],1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A5'

wb.save('DangKyCa-CoSo3.xlsx')
print('SAVED DangKyCa-CoSo3.xlsx  cols:', last_col, 'rows:', LR)
